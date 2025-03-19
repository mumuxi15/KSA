# ksa_api/pricing_handler.py
import pandas as pd
import os
import urllib3
import requests
import openpyxl
from io import BytesIO
from openpyxl.styles import Border, Side, Font
from openpyxl.drawing.image import Image
from .config import *
from outlook.email_handler import request_pricing, get_employee_emails
from query.db_handler import query_search, query_direct
from .utils import get_inactive_items, get_reviewed_list, format_table, get_supplier_emails

# Thin border style for Excel formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

class PricingHandler:
    def __init__(self):
        self.gsheet_key = GSHEET_KEY
        self.pricing_worksheet_key = PRICING_WORKSHEET_KEY
        self.prc_folder = PRICING_PATH["prc_folder"]
        self.data_folder = PRICING_PATH["data_folder"]
        self.sample_image_url = SAMPLE_IMAGE_URL

    def get_inactive_items(self):
        return get_inactive_items(self.gsheet_key, self.pricing_worksheet_key)

    def get_reviewed_list(self):
        return get_reviewed_list(self.gsheet_key, self.pricing_worksheet_key)

    def get_vendor_items(self, vendor):
        df = query_search(target='vendoritems', item_tuple=f"('{vendor}')")
        if not df['COST'].equals(df['COST_FOR']):
            rates = round((df['COST_FOR'] / df['COST']).mean(), 1)
            df['CURR'] = FX_RATES[rates]
        return df

    def request_vendor_price(self, year=2024):
        """Request pricing from vendors for a given year."""
        tables = query_search(target='uprice')
        df_supplr = get_supplier_emails(self.gsheet_key, self.pricing_worksheet_key)

        # Data filtering
        tables['SELSTA'] = tables['SELSTA'].fillna('')
        tables = tables.loc[tables['Season'] <= str(year)[-2:]]  # Only get current season
        tables.loc[(tables['SELSTA'] == 'SO-NR') & (tables['QOH'] <= 0), 'STATUS'] = 'OUT'

        inactive_vendors = self.get_inactive_items()
        tables = tables.loc[~tables['SUPPLR'].isin(inactive_vendors)]
        tables = tables.merge(df_supplr, on='SUPPLR', how='inner')  # Filter selective suppliers only

        tables = tables.rename(columns={'FirstCost': f'{year}Cost($)', 'NewCost': f'{year + 1}Cost'})
        no_img_available = BytesIO(requests.get(self.sample_image_url).content)  # No image available placeholder

        for vendor in sorted(tables['SUPPLR'].unique()):
            vendoremail = df_supplr.loc[df_supplr['SUPPLR'] == vendor, 'EMAIL'].values[0]
            vendorname = df_supplr.loc[df_supplr['SUPPLR'] == vendor, 'Supplier'].values[0]
            df = tables[tables['SUPPLR'] == vendor].copy()

            if not df[f'{year}Cost($)'].equals(df['COST_FOR']):
                rates = round((df['COST_FOR'] / df[f'{year}Cost($)']).mean(), 1)
                curr = FX_RATES[rates]
                df[f'{year}Cost($)'] = df['COST_FOR']
                df = df.rename(columns={f'{year}Cost($)': f'{year}Cost({FX_SYMBOL[curr]})'})

            df = df.drop(columns=['SELSTA', 'QOH', 'Season', 'Supplier', 'COST_FOR', 'EMAIL']).sort_values(by=['PRODCT']).reset_index(drop=True)
            output_file = f'{self.prc_folder}KSA 2025 PRICING {vendor}.xlsx'
            df.to_excel(output_file, startcol=0, startrow=3)

            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            ws['A1'] = 'Exchange Rate: '
            ws['A2'] = 'Country: '
            ws['F1'] = f'Name: {vendorname}'
            ws['F2'] = f'Code#: {vendor}'

            # Add border to each cell
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 9:  # Column I
                        ws[f'H{cell.row}'] = f'=G{cell.row}/F{cell.row}-1'

            # Insert images if fewer than 100 items
            if len(df) < 99:
                df['Photo'] = df['PRODCT'].str.lower().apply(
                    lambda x: f'https://www.ksa.com/media/catalog/product/{x[0]}/{x[1]}/{x}.jpg')
                http = urllib3.PoolManager()
                for idx, row in df.iterrows():
                    ws.row_dimensions[idx + 5].height = 50  # Start on row 4
                    response = http.request('GET', f'{row["Photo"]}?width=50')
                    img = BytesIO(response.data)  # Byte form
                    img.width, img.height = 50, 50
                    if img.getvalue() != no_img_available.getvalue():
                        ws.add_image(Image(img), f'D{idx + 5}')

            wb.save(output_file)
            request_pricing(recipient=vendoremail, vendor=vendor, attachment=output_file)
            print('--' * 10)

    def gen_cover_sheets(self, includes=[], new_year=2025, compare_crystal=False):
        """Generate cover sheets for vendors."""
        cost_col = f'{new_year}Cost'
        exchange_rates = self.get_reviewed_list()['Exchange Rate']
        reviewed_vendors = exchange_rates.keys()
        covered = [f.split('_')[0] for f in os.listdir(f'{self.prc_folder}COVERS/')]
        discrepancy = list(set(reviewed_vendors) - set(covered))
        print('PROCESSING . . . ', discrepancy)

        files = [f for f in os.listdir(self.prc_folder) if f.endswith('.xls') or f.endswith('.xlsx')]
        files = [f for f in files if f.split('.')[0].split(' ')[3] in discrepancy]
        report = None # report = read_crystal_report()
        for f in files:
            vendor = f.split('.')[0].split(' ')[3]
            price = pd.read_excel(f'{self.prc_folder}{f}', skiprows=3)
            price = price.rename(columns={
                **{c: 'PRODCT' for c in price.columns if ('Item' in c) or ('ITEM' in c)},
                **{c: f'{new_year}Cost' for c in price.columns if str(new_year) in c},
                **{c: 'MOQ' for c in price.columns if ('MIN' in c) or ('MIN QTY' in c)}
            })

            if compare_crystal:
                df = self.check_wt_cr01(vendor=vendor, crystal=report.loc[report['SUPPLR'] == vendor, 'PRODCT'])
            else:
                df = self.get_vendor_items(vendor)
            df = df.merge(price[['PRODCT', cost_col, 'MOQ']], how="left", on="PRODCT", indicator=True)
            vendor_longname = df['Supplier'].unique()[0]
            df['%Change'] = round((df[cost_col] / df['COST_FOR'] - 1) * 100, 1).fillna(0)
            if df['%Change'].mean() == 0:
                price_pct = 'no price change'

            ####  check for duplicated products, customer might offer QTY discount, diff Price for diff MOQ
            if not df[df['PRODCT'].duplicated()].empty:
                print('YOU NEED TO CODE HERE ')
                print(df[df['PRODCT'].duplicated()])
                print('-!!!!!!!!-' * 2)
                df['sorting'] = df['MOQ'].apply(lambda x: 1 if x == 0 else 0)
                df = df.sort_values(by=['sorting', 'PRODCT'])
            else:
                df = df.drop(columns=['MOQ'])
            ###  1.check if item is OUT or not
            ###  2.if item is nan, a) UNQ b)SO-NR - do not bother asking, just use old price
            df[cost_col] = df[cost_col].apply(lambda x: '{:,.2f}'.format(x))
            df.loc[((df['STATUS'] == 'P') | (df['STATUS'] == 'PP')) & (df['QOH'] <= 0), cost_col] = 'OUT'
            df.loc[(df[cost_col] == 'nan') & (df['SELSTA'] == 'UNQ'), cost_col] = 'UNQ'
            df.loc[(df[cost_col] == 'nan') & (df['QOH'] > 0) & (
                        (df['STATUS'] == 'P') | (df['STATUS'] == 'PP')), cost_col] = df['COST_FOR'].apply(
                lambda x: '{:,.2f}'.format(x))

            df.loc[df['SELSTA'] == 'SO-NR', '_merge'] = 'both'
            df.loc[df[cost_col] == 'UNQ', '_merge'] = 'both'
            error = df[df['_merge'] != 'both']

            epsilon = df.loc[df[cost_col] == 'nan']
            if not epsilon.empty:
                print('---  NAN ERROR    -----')
                print(epsilon, '\n', '----' * 5)
                flag += '_nan_error'

            if len(error) != 0:
                print(vendor, ' EXCEL SHEET ISSUE  \n', error)
                flag += "_supplr_flag"
            else:
                df = df.drop(columns=['_merge'])

            df = df.drop(columns=['Supplier', 'Season', 'COST', 'COST_FOR'])
            df = df.rename(columns={'Item': 'PRODCT'}).reset_index(drop=True)
            cost_col_index = df.columns.get_loc('2025Cost')
            #   FORMATING
            # return
            writer = pd.ExcelWriter(f'{PATH['prc_folder']}COVERS/{vendor}_cover_sheet{flag}.xlsx')
            df.to_excel(writer, startcol=0, startrow=1)
            #
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            border_format = workbook.add_format({'border': 1})
            format = workbook.add_format({'font_size': 10})
            format_red = workbook.add_format({'font_color': 'red'})
            # Apply the border format to the entire table
            worksheet.conditional_format(f'A2:{chr(ord('A') + len(df.columns))}{len(df) + 2}',
                                         {'type': 'no_errors', 'format': border_format})
            worksheet.conditional_format(2, cost_col_index, len(df) + 2, cost_col_index + 1,
                                         {'type': 'cell', 'criteria': '==', 'value': '"OUT"', 'format': format_red})
            worksheet.set_column('A:Z', None, format)
            worksheet.set_column('F:F', 13)

            worksheet.set_landscape()
            for i in range(df.shape[0] + 2):  # +1 to include the header row
                worksheet.set_row(i, 12)
            worksheet.write_string(0, 0, exchange_rates[vendor])
            worksheet.write_string(0, 1, vendor_longname)
            worksheet.write_string(0, 9, price_pct)
            writer._save()

        return

    def check_system_prices(self):
        filenames = os.listdir(self.prc_folder + '/COVERS/COMPLETED')
        vendors = [f.split('_')[0] for f in filenames]
        df = []
        for f in filenames:
            df.append(pd.read_excel(self.prc_folder + '/COVERS/COMPLETED/' + f, skiprows=1))

        df = pd.concat(df)[['ITEST', 'SELSTA', 'SUPPLR', 'PRODCT', '2025Cost', 'CURR', 'Chg']]
        prices = query_search(target='uprice')[['Item', 'FirstCost']]
        df = df.merge(prices, how="outer", left_on="PRODCT", right_on="Item")
        df['FirstCost'] = df['FirstCost'].apply(lambda x: "{:.2f}".format(x))
        df['flag'] = (df['FirstCost'] != df['2025Cost']) * 1
        df.loc[df['2025Cost'] == 'OUT', 'flag'] = 0
        print(df[df['flag'] == 1])

    def count_vendor_new_items():
        data = query_search(target="newbyyear")
        # no filter on item status
        data['date'] = data['date'].dt.year
        data = data.loc[data['date'] >= 2020]
        correction = {'J & J SEASONAL COMPANY LTD. (H03902)': 'J&J SEASONAL COMPANY LTD. (H03902)'}
        data["Longname"] = data["Longname"].replace(correction)
        data['primary'] = data['Supplier'].str[0:4]
        data['shortname'] = data['Longname'].apply(lambda x: x.split()[0])
        supp = data[['Supplier', 'Longname', 'primary', 'shortname']].drop_duplicates()

        data = data.groupby(['Supplier', 'date']).agg({'Item': 'count'}).reset_index()
        for origin in ['C', 'H', 'T']:
            df = data.loc[data['Supplier'].str[0] == origin]
            df = pd.pivot_table(df, values='Item', columns='date', index='Supplier')
            df = supp.merge(df, on=['Supplier'], how='right')
            df = df.groupby(['primary', 'shortname']).sum().reset_index()
            df = df.drop(columns=['shortname'])
            df['Longname'] = df['Longname'].apply(lambda x: ' '.join(x.split()[:3]))
            df = df.sort_values(by=['primary', 'Supplier'])
            df.to_csv(f'{origin}_vendors.csv')
            print(df)
