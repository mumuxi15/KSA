import gspread
import pandas as pd
import os, urllib3
import requests
import openpyxl
from io import BytesIO
from openpyxl.styles import NamedStyle,  Border, Side,Font
from openpyxl.drawing.image import Image
from sqlalchemy import exc
from query import query_search, query_direct
from config.env import env, PATH, FX_RATES, FX_SYMBOL
from shipping import send_pricing_email, get_employee_emails

pd.set_option('display.max_columns',20)
pd.set_option('display.width', 200)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def test_gsheet_connection():
    """ TEST CONNECTION, LIST ALL AVAILABLE SHEETS  """
    gc = gspread.service_account(filename=PATH['gsheet_key'])
    spreadsheets = gc.openall()
    if spreadsheets:
        for spreadsheet in spreadsheets:
            print("Title:", spreadsheet.title, "URL:", spreadsheet.url)
    else:
        print("No spreadsheets available, Please share the spreadsheet with Service Account email")

def gc_inactive_suppliers() -> list:
    gc = gspread.service_account(filename=PATH['gsheet_key'])
    sh = gc.open_by_key(env("PRICING_WORKSHEET_KEY"))
    df = pd.DataFrame(sh.get_worksheet(1).get_all_records()) # 2nd sheet in line
    return df.loc[df['OUT'] == 'x', 'SUPPLR'].to_list()

def get_supplier_emails():
    """ GET SUPPLIER EMAIL FROM GSHEET  """
    gc = gspread.service_account(filename=PATH['gsheet_key'])
    sh = gc.open_by_key(env("PRICING_WORKSHEET_KEY"))
    df = pd.DataFrame(sh.get_worksheet(0).get_all_records())
    # d2 = pd.DataFrame(sh.get_worksheet(2).get_all_records())
    df = df.loc[df['SENT']=='']
    return df[['SUPPLR','EMAIL']]

def get_inactive_items():
    """ GET LIST OF INACTIVE ITEMS FROM GSHEET """
    gc = gspread.service_account(filename=PATH['gsheet_key'])
    sh = gc.open_by_key(env("PRICING_WORKSHEET_KEY"))
    df = pd.DataFrame(sh.get_worksheet(3).get_all_records())
    df = df.loc[~(df['Active']==1)]
    return df['Item'].to_list()

def get_reviewed_list():
    """GET LIST OF VENDORS THAT COMPLETED THE TASK """
    gc = gspread.service_account(filename=PATH['gsheet_key'])
    sh = gc.open_by_key(env("PRICING_WORKSHEET_KEY"))
    df = pd.DataFrame(sh.get_worksheet(0).get_all_records())[['SUPPLR','REVIEWED','PRINTED', 'Exchange Rate']]
    df = df.loc[(df['PRINTED']!=1)&((df['REVIEWED']==1)|(df['REVIEWED']==2)),['SUPPLR', 'Exchange Rate']]
    df.index=df['SUPPLR']
    df = df[['Exchange Rate']]
    return df.to_dict()

def get_inactive_supplier(data, year = 2024, create_covers=False):
    """vendors with all items UNQ and list from google spreadsheet
       return list of inactive suppliers;
       option: create cover sheets
    """
    ds = data.groupby('SUPPLR').agg({'PRODCT':lambda x: x.nunique(), 'SELSTA': lambda x: (x=='UNQ').sum()})
    ds = ds[ds['PRODCT']==ds['SELSTA']]
    inactive_vendors = set(gc_inactive_suppliers())|(set(ds.index))
    if create_covers:
        for vendor in inactive_vendors:
            # print (vendor)
            price_col = f'{year+1}Cost'
            df = data.loc[data['SUPPLR']==vendor].copy()
            vendorname = df['Supplier'].unique()[0]
            df['FirstCost'] = df['FirstCost'].apply(lambda x: '{:,.2f}'.format(x))
            df.loc[df['SELSTA'] == 'UNQ', 'FirstCost'] = 'UNQ'
            df.loc[(df['ITEST'] != 'R') & (df['QOH'] <= 0), 'FirstCost'] = 'OUT'
            df = df.rename(columns={'FirstCost':price_col})
            df = df[['ITEST', 'SELSTA', 'QOH', 'SUPPLR', 'STATUS', 'PRODCT', price_col]].reset_index(drop=True).sort_values(by='PRODCT')

            output_file = f'{PATH['prc_folder']}COVERS/{vendor}_cover_sheet2.xlsx'
            df.to_excel(output_file, startcol=0, startrow=1)
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            ws['B1'] = vendorname
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.value == 'OUT':
                        cell.font = Font(color="FF0000")
            wb.save(output_file)
    return inactive_vendors

def request_vendor_price(year=2024 ):
    tables = query_search(target='uprice')
    df_supplr = get_supplier_emails()
    # data filter   1. !=UNQ  2. only supplr sent==''
    tables['SELSTA'] = tables['SELSTA'].fillna('')
    tables = tables.loc[tables['Season']<=str(year)[-2::]]  #only get current season
    tables.loc[(tables['SELSTA']=='SO-NR')&(tables['QOH']<=0),'STATUS'] = 'OUT'

    inactive_vendors = get_inactive_supplier(tables, create_covers=False)
    tables = tables.loc[~(tables['SUPPLR'].isin(inactive_vendors))]
    ## google sheet sheet 1: summary table
    # tables[['SUPPLR']].drop_duplicates().sort_values(by='SUPPLR').to_csv('summary_all_supplier.csv')
    tables = tables.merge(df_supplr, on='SUPPLR', how='inner')  # filtering selective suppler only
    df_supplr = tables[['SUPPLR','Supplier','EMAIL']].drop_duplicates()
    if not tables.loc[tables['Season']!=str(year)[-2::]].empty:
        print ('potential error with those products ')
        print (tables.loc[tables['Season']!=str(year)[-2::]]) #items

    tables = tables.rename(columns={'FirstCost': f'{year}Cost($)','NewCost':f'{year+1}Cost'})
    no_img_available = BytesIO(requests.get(PATH['sample_image']).content)  # no image available array

    for vendor in sorted(tables['SUPPLR'].unique()):
        "if more than 100 items, item images not required"
        vendoremail = df_supplr.loc[df_supplr['SUPPLR']==vendor,'EMAIL'].values[0]
        vendorname = df_supplr.loc[df_supplr['SUPPLR']==vendor,'Supplier'].values[0]
        df = tables[tables['SUPPLR']==vendor].copy()
        if not df[f'{year}Cost($)'].equals(df['COST_FOR']):
            rates = round((df['COST_FOR'] / df[f'{year}Cost($)']).mean(), 1)
            curr = FX_RATES[rates]
            df[f'{year}Cost($)'] = df['COST_FOR']
            df = df.rename(columns={f'{year}Cost($)':f'{year}Cost({FX_SYMBOL[curr]})'})

        df = df.drop(columns=['SELSTA', 'QOH', 'Season', 'Supplier', 'COST_FOR', 'EMAIL']).sort_values(by=['PRODCT']).reset_index(drop=True)
        print (df)
        output_file = f'{PATH['prc_folder']}KSA 2025 PRICING {vendor}.xlsx'
        df.to_excel(output_file, startcol=0, startrow=3)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        ws['A1'] = 'Exchange Rate: '
        ws['A2'] = 'Country: '
        ws['F1'] = f'Name: {vendorname}'
        ws['F2'] = f'Code#: {vendor}'
        # Add border to each cell
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.column == 9: # col I = 9
                    ws[f'H{cell.row}'] = f'=G{cell.row}/F{cell.row}-1'
        ### Insert Image
        if len(df)<99:
            df['Photo'] = df['PRODCT'].str.lower().apply(lambda x: f'https://www.kurtadler.com/media/catalog/product/{x[0]}/{x[1]}/{x}.jpg')
            http = urllib3.PoolManager()
            for idx, row in df.iterrows():
                ws.row_dimensions[idx + 5].height = 50  # start on row 4
                response = http.request('GET', f'{row['Photo']}?width=50')
                img = BytesIO(response.data)  #byte form
                img.width, img.height = 50, 50
                if img.getvalue() != no_img_available.getvalue():
                    ws.add_image(Image(img), f'D{idx+5}')
        else:
            print ('TOO BIG')

        wb.save(output_file)
        # send_pricing_email(recipient=get_employee_emails('PP'), vendor=vendor,attachment=output_file)
        # send_pricing_email(recipient=vendoremail,vendor=vendor, attachment=output_file)
        print ('--'*10)
    return
def get_vendor_items(vendor):
    df = query_search(target='vendoritems', item_tuple=f"('{vendor}')")
    if not df['COST'].equals(df['COST_FOR']):
        rates = round((df['COST_FOR'] / df['COST']).mean(), 1)
        df['CURR'] = FX_RATES[rates]
    return df

def check_wt_cr01(vendor, crystal, year=2025):
    df = get_vendor_items(vendor)
    # MERGE DATABASE DATA WITH CRYSTAL REPORT EXCEL FILE
    df = df.merge(crystal, on='PRODCT', how='outer', indicator=True)
    error = df[df['_merge'] != 'both']
    if not error.empty:
        print('----- CHECK WITH CR01 ----',error,'\n ','++'*20)
        if error['Season'].unique()[0] != str(year)[-2::]:
            print ('PREVIOUS SEASON FOUND ! ')
            if error.loc[error['ITEST']=='R'].empty:
                'CAN SAFELY IGNORE THE ERROR'
                df = df.rename(columns={'_merge': 'data_issue'})
                df['data_issue'] = df['data_issue'].str.replace('both', '')
                return df
            else:
                raise ValueError('RETURN item found in previous seasons')
        else:
            raise ValueError('go to line 189 and see what issue you have ')
    else:
        df = df.drop(columns=['_merge'])
        return df

def gen_cover_sheets(includes=[],new_year=2025, compare_crystal=False):
    """ GENERATE COVER SHEETS """
    cost_col = f'{new_year}Cost'
    exchange_rates = get_reviewed_list()['Exchange Rate']
    reviewed_vendors = exchange_rates.keys()
    covered = [f.split('_')[0] for f in os.listdir(PATH['prc_folder']+'/COVERS/')]
    discrepancy = list(set(reviewed_vendors) - set(covered))
    print ('PROCESSING . . . ',discrepancy)
    files = [f for f in os.listdir(PATH['prc_folder']) if f.endswith('.xls') or f.endswith('.xlsx')]
    files = [f for f in files if f.split('.')[0].split(' ')[3] in discrepancy]
    report = read_crystal_report()
    for f in files:
        vendor = f.split('.')[0].split(' ')[3]
        flag, price_pct = "", ""
        price = pd.read_excel(PATH['prc_folder']+f, skiprows=3)
        price = price.rename(columns={
            **{c:'PRODCT' for c in price.columns if ('Item' in c) or ('ITEM' in c)},
            **{c: cost_col for c in price.columns if str(new_year) in c},
            **{c:'MOQ' for c in price.columns if ('MIN' in c) or ('MIN QTY' in c)}
        })  # find the item column and rename to "Item"
        ## 1. pull all items from db where supprl=vendor
        #  2. check with CR01 report
        #  3. merge with pricing sheet ####
        # print (dp.head())
        if compare_crystal:
            df = check_wt_cr01(vendor=vendor, crystal=report.loc[report['SUPPLR'] == vendor, 'PRODCT'])
        else:
            df = get_vendor_items(vendor)
        df = df.merge(price[['PRODCT',cost_col,'MOQ']], how="left", on="PRODCT", indicator=True)
        vendor_longname = df['Supplier'].unique()[0]
        df['%Change'] = round((df[cost_col]/df['COST_FOR']-1)*100,1).fillna(0)
        if df['%Change'].mean() == 0:
            price_pct = 'no price change'

        ####  check for duplicated products, customer might offer QTY discount, diff Price for diff MOQ
        if not df[df['PRODCT'].duplicated()].empty:
            print ('YOU NEED TO CODE HERE ')
            print (df[df['PRODCT'].duplicated()])
            print ('-!!!!!!!!-'*2)
            df['sorting'] = df['MOQ'].apply(lambda x: 1 if x==0 else 0)
            df = df.sort_values(by=['sorting', 'PRODCT'])
        else:
            df = df.drop(columns=['MOQ'])
        ###  1.check if item is OUT or not
        ###  2.if item is nan, a) UNQ b)SO-NR - do not bother asking, just use old price
        df[cost_col] = df[cost_col].apply(lambda x: '{:,.2f}'.format(x))
        df.loc[((df['STATUS']=='P')|(df['STATUS']=='PP')) & (df['QOH']<=0),cost_col] = 'OUT'
        df.loc[(df[cost_col]=='nan')&(df['SELSTA']=='UNQ'), cost_col] = 'UNQ'
        df.loc[(df[cost_col]=='nan')&(df['QOH']>0)&((df['STATUS']=='P')|(df['STATUS']=='PP')), cost_col] = df['COST_FOR'].apply(lambda x: '{:,.2f}'.format(x))


        df.loc[df['SELSTA']=='SO-NR','_merge']='both'
        df.loc[df[cost_col]=='UNQ','_merge']='both'
        error = df[df['_merge']!='both']


        epsilon = df.loc[df[cost_col]=='nan']
        if not epsilon.empty:
            print ('---  NAN ERROR    -----')
            print (epsilon,'\n','----'*5)
            flag += '_nan_error'

        if len(error)!=0:
            print (vendor, ' EXCEL SHEET ISSUE  \n',error)
            flag += "_supplr_flag"
        else:
            df = df.drop(columns=['_merge'])

        df = df.drop(columns=['Supplier','Season','COST','COST_FOR'])
        df = df.rename(columns={'Item':'PRODCT'}).reset_index(drop=True)
        cost_col_index = df.columns.get_loc('2025Cost')
        #   FORMATING
        # return
        writer = pd.ExcelWriter(f'{PATH['prc_folder']}COVERS/{vendor}_cover_sheet{flag}.xlsx')
        df.to_excel(writer, startcol=0, startrow=1)
        #
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        border_format = workbook.add_format({'border': 1})
        format = workbook.add_format({'font_size': 10})
        format_red = workbook.add_format({'font_color': 'red'})
        # Apply the border format to the entire table
        worksheet.conditional_format(f'A2:{chr(ord('A') + len(df.columns))}{len(df) + 2}',
                                     {'type': 'no_errors', 'format': border_format})
        worksheet.conditional_format(2, cost_col_index, len(df)+2, cost_col_index+1,{'type': 'cell', 'criteria': '==','value': '"OUT"','format': format_red})
        worksheet.set_column('A:Z', None, format)
        worksheet.set_column('F:F', 13)

        worksheet.set_landscape()
        for i in range(df.shape[0] + 2):  # +1 to include the header row
            worksheet.set_row(i, 12)
        worksheet.write_string(0, 0, exchange_rates[vendor])
        worksheet.write_string(0, 1, vendor_longname)
        worksheet.write_string(0, 9, price_pct)
        writer._save()


def read_crystal_report():
    df = pd.read_excel(PATH['data_folder']+'victor old landed sheets.xlsx',skiprows=1)
    col = list(df.columns)
    i = col.index('PRODCT')
    df = df[col[i-1:i+1]]
    df['PRODCT'] = df['PRODCT'].fillna(df[col[i-1]])
    df = df['PRODCT'].str.split(' ',expand=True)
    df = df[[0]]
    df.columns=['Item']
    df = df.dropna(how="all")
    df = df.loc[~df['Item'].str.contains('PRODCT')]
    df = df.loc[df['Item']!='OV0001']
    df = query_search(target="getsupplr", item_tuple=tuple(df['Item'].values))
    return df

def test_database():
    # from query import query_direct
    tb = """
"""
    tbs = [t for t in tb.split('\n') if t!=""]
    for tb in tbs:
        try:
            df = query_search(target="testdb", item_tuple=tb)
            if len(df)<30:
                print (tb, '   SKIP  ')
            else:
                print (df.columns)
                suffix = ''
                if 'TRANSCTN' in df.columns:
                    suffix='_transctn'
                elif 'PRODCT' in df.columns:
                    suffix='_prod'
                elif 'INVOICE' in df.columns:
                    suffix='_invoice'

                df.to_csv(tb+suffix+'.csv')
        except exc.SQLAlchemyError:
            print (tb, '  db issues  SKIP  ')

def rename_files():
    import os
    path = f'{PATH['prc_folder']}/COMPLETED/'
    filenames = os.listdir(path)
    filenames = [f for f in filenames if 'PRICNG' in f]
    print(filenames)
    for filename in filenames:
        newname = os.path.join(path,filename.replace("PRICNG", "PRICING"))
        oldname = os.path.join(path,filename)
        os.rename(oldname,newname)

def serena_files():
    # re-edit her sheets
    path = f'{PATH['prc_folder']}serena/'
    filenames = os.listdir(path)
    for filename in filenames:
        vendor = filename.split('.')[0].split(' ')[-1]
        print (vendor)
        df = pd.read_excel(f'{PATH['prc_folder']}serena/{filename}')
        s = df[df.columns[0]]
        i = s[s.str.contains('KSA', na=False)].index[0]
        df = pd.read_excel(f'{PATH['prc_folder']}serena/{filename}',skiprows=i+1)
        col2025 = [c for c in df.columns if '2025' in c][0]
        col2024 = [c for c in df.columns if '2024' in c][0]
        df = df.rename(columns={col2025:"2025 COST", col2024: "2024 COST", "NW/GW LBS":"NW/GW","2024 PURCH":"PURCH"})
        df = df.dropna(how="all",axis=1)
        df['DIMENSIONS (INCHES)'] =  df['DIMENSIONS (INCHES)'].str.replace(r'\s+', ' ', regex=True).str.replace('CTN','\nCTN')
        df = df.drop(columns=['IN/OUT/   PACK PCS',])
        df['Photo'] =''
        df['% Change'] = df['2025 COST']/df['2024 COST']-1
        df = df[list(df.columns[0:1])+['Photo']+list(df.columns[1:4])+['% Change']+list(df.columns[4:-2])]

        writer = pd.ExcelWriter(f'{PATH['prc_folder']}/KSA 2025 PRICING {vendor}.xlsx')
        df.to_excel(writer, startcol=0, startrow=3)
        #
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        border_format = workbook.add_format({'border': 1})

        # Apply the border format to the entire table
        worksheet.conditional_format(f'A4:{chr(ord('A') + len(df.columns))}{len(df) + i-1}',
                                     {'type': 'no_errors', 'format': border_format})

        format1 = workbook.add_format({'num_format': '0%'})
        # worksheet.set_column('C:C', None, format2)
        worksheet.set_column('D:D', 20)
        worksheet.set_column('L:L', 18)
        for r in range(4,len(df)+5):
            worksheet.set_row(r,height=49)

        worksheet.write_string(0, 0, 'EXCHANGE RATE:___RMB 7.00___')
        worksheet.write_string(1, 0, 'PRICES QUOTED DO NOT INCLUDE UPC BAR CODING')
        worksheet.write_string(0, 5, f'SUPPLIER:  ')
        worksheet.write_string(1, 5, f'VENDOR#:')
        worksheet.write_string(1, 6, vendor)
        worksheet.set_landscape()

        writer._save()

        ### add currency

        wb = openpyxl.load_workbook(f'{PATH['prc_folder']}/KSA 2025 PRICING {vendor}.xlsx')
        ws = wb.active
        currency_style = NamedStyle(name='currency_rmb', number_format='¥#,##0.00')
        # Apply the currency style to columns E and F
        for col in ['E', 'F']:
            for row in ws[col]:
                row.number_format = currency_style.number_format
        # Save the workbook with currency formatting
        wb.save(f'{PATH['prc_folder']}/KSA 2025 PRICING {vendor}.xlsx')
    return
def check_system_prices():
    filenames = os.listdir(PATH['prc_folder'] + '/COVERS/COMPLETED')
    vendors = [f.split('_')[0] for f in filenames]
    df = []
    for f in filenames:
        df.append(pd.read_excel(PATH['prc_folder'] + '/COVERS/COMPLETED/' + f, skiprows=1))

    df = pd.concat(df)[['ITEST', 'SELSTA', 'SUPPLR', 'PRODCT', '2025Cost', 'CURR', 'Chg']]
    prices = query_search(target='uprice')[['Item','FirstCost']]
    df = df.merge(prices,how="outer", left_on="PRODCT", right_on="Item")
    df['FirstCost'] = df['FirstCost'].apply(lambda x: "{:.2f}".format(x))
    df['flag'] = (df['FirstCost']!=df['2025Cost'])*1
    df.loc[df['2025Cost']=='OUT','flag'] = 0
    print (df[df['flag']==1])
def count_vendor_new_items():
    data = query_search(target="newbyyear")
    # no filter on item status
    data['date'] = data['date'].dt.year
    data = data.loc[data['date']>=2020]
    correction = {'J & J SEASONAL COMPANY LTD. (H03902)':'J&J SEASONAL COMPANY LTD. (H03902)'}
    data["Longname"] = data["Longname"].replace(correction)
    data['primary'] = data['Supplier'].str[0:4]
    data['shortname'] = data['Longname'].apply(lambda x: x.split()[0])
    supp = data[['Supplier','Longname','primary','shortname']].drop_duplicates()

    data = data.groupby(['Supplier','date']).agg({'Item':'count'}).reset_index()
    for origin in ['C','H','T']:
        df = data.loc[data['Supplier'].str[0]==origin]
        df = pd.pivot_table(df, values='Item',columns='date',index='Supplier')
        df = supp.merge(df,on=['Supplier'],how='right')
        df = df.groupby(['primary','shortname']).sum().reset_index()
        df = df.drop(columns=['shortname'])
        df['Longname'] = df['Longname'].apply(lambda x: ' '.join(x.split()[:3]))
        df = df.sort_values(by=['primary','Supplier'])
        df.to_csv(f'{origin}_vendors.csv')
        print (df)

def send_to_printer():
    import win32api
    import win32print
    import time
    # all_printers = win32print.EnumPrinters(2)
    defaultPrinter = win32print.GetDefaultPrinter()
    print (defaultPrinter)
    # if defaultPrinter != 'Art Dept C8100-716E PS':
    win32print.SetDefaultPrinter('Art Dept C8100-716E PS')
    mydir = "C:/Users/mpan/Desktop/print"
    files = os.listdir(mydir)
    for f in files:
        print ( "printing file " + str(mydir + f) + " on " + str(win32print.GetDefaultPrinter()))
        win32api.ShellExecute(0, "print", os.path.join(mydir, f), None, ".", 0)
        time.sleep(6)

def test():
    df = query_direct("SELECT  * FROM kuradl_97x_740_prod_x.dbo.ICPC")
    print (df)
    # df.to_csv('meng.csv')
    return


def main():
    test()
    # read_new_item_info_sheet()
    # request_vendor_price()
    # gen_cover_sheets(includes=['MOQ'])
    # gen_cover_sheets()
    # gen_cover_sheets_inactive()
    # read_crystal_report()
    # rename_files()
    # serena_files()
    # check_system_prices()
    # update_all_new_prices()
    # count_vendor_new_items()
    # test_database()
    # send_to_printer()

main()